from __future__ import annotations
from io import BytesIO
from openpyxl import load_workbook


def _val(row, idx: int) -> str:
    return str(row[idx]).strip() if idx < len(row) and row[idx] is not None else ""


# Labels in col I (MUTATIE column) that mark sub-sections — not detainee names
_SECTION_LABELS = {
    "pac", "ziekenhuis", "straf en veiligheidscellen", "naam gedetineerde", "mutatie",
}


def parse_werkers_mutaties(file_bytes: bytes) -> dict:
    """
    Parse the MUTATIES sheet from a filled mutatielijst.

    Sheet layout (row 3 = header, data from row 4):
      Col A(0): CEL incoming  Col B(1): INKOMENDE  Col C(2): BK
      Col E(4): CEL outgoing  Col F(5): UITGAANDEN  Col G(6): REDEN
      Col I(8): MUTATIE naam  Col J(9): VAN cel     Col K(10): NAAR cel

    Returns:
      {
        "uitgaanden": [{"cel": ..., "naam": ..., "reden": ...}, ...],
        "mutaties":   [{"naam": ..., "van": ..., "naar": ...}, ...],
      }
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    if "MUTATIES" not in wb.sheetnames:
        return {"uitgaanden": [], "mutaties": []}

    ws = wb["MUTATIES"]
    result: dict = {"uitgaanden": [], "mutaties": []}

    # Locate header row (contains both UITGAANDEN and MUTATIE)
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        vals = [str(c).upper().strip() if c else "" for c in row]
        if "UITGAANDEN" in vals and "MUTATIE" in vals:
            header_row = i
            break

    if header_row is None:
        return result

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        if all(v is None for v in row):
            continue

        # ── UITGAANDEN ────────────────────────────────────────────────────────
        naam_uit = _val(row, 5)
        if naam_uit:
            result["uitgaanden"].append({
                "cel":   row[4],
                "naam":  naam_uit,
                "reden": _val(row, 6),
            })

        # ── MUTATIE ───────────────────────────────────────────────────────────
        naam_mut = _val(row, 8)
        if naam_mut and naam_mut.lower() not in _SECTION_LABELS:
            van  = row[9]
            naar = row[10]
            if van is not None and naar is not None:
                result["mutaties"].append({
                    "naam": naam_mut,
                    "van":  van,
                    "naar": naar,
                })

    return result
