from __future__ import annotations
from io import BytesIO
from openpyxl import load_workbook
from .normalizer import normalize_name, normalize_cell, normalize_key


def parse_celbezetting(file_bytes: bytes) -> dict:
    """
    Returns dict keyed on normalize_key(naam, voornaam) → DetaineeRecord dict.
    Also returns a list for autocomplete: [{"naam", "voornaam", "cel", "key"}]
    """
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Detect real header row (contains Naam, Voornaam, Cel)
    header_row_idx = None
    col_map = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
        if "naam" in row_lower and "voornaam" in row_lower and "cel" in row_lower:
            header_row_idx = i
            for j, val in enumerate(row):
                if val is not None:
                    col_map[str(val).strip().lower()] = j
            break

    if header_row_idx is None:
        raise ValueError("Geen header-rij gevonden in celbezetting (verwacht: Naam, Voornaam, Cel)")

    idx_naam = col_map.get("naam", 0)
    idx_voor = col_map.get("voornaam", 1)
    idx_cel = col_map.get("cel", 4)
    idx_tijdcel = col_map.get("tijdelijke cel", 7)
    idx_exstat = col_map.get("ext.stat.", 6)

    lookup: dict[str, dict] = {}
    autocomplete: list[dict] = []

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        naam_raw = row[idx_naam] if len(row) > idx_naam else None
        voor_raw = row[idx_voor] if len(row) > idx_voor else None
        cel_raw = row[idx_cel] if len(row) > idx_cel else None
        tijdcel_raw = row[idx_tijdcel] if len(row) > idx_tijdcel else None

        if not naam_raw:
            continue

        naam_str = str(naam_raw).strip()
        voor_str = str(voor_raw).strip() if voor_raw else ""
        cel_int = normalize_cell(cel_raw)
        tijdcel_int = normalize_cell(tijdcel_raw)

        key = normalize_key(naam_str, voor_str)
        record = {
            "naam": naam_str,
            "voornaam": voor_str,
            "cel": cel_int,
            "tijdelijke_cel": tijdcel_int,
            "key": key,
        }
        lookup[key] = record
        autocomplete.append(record)

    return {"lookup": lookup, "autocomplete": autocomplete}
