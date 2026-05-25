from __future__ import annotations
from io import BytesIO
from datetime import datetime, time
from openpyxl import load_workbook
from .normalizer import normalize_cell


def _to_time(val) -> time | None:
    if val is None:
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    s = str(val).strip()
    if not s or s in ("\xa0", ""):
        return None
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            pass
    return None


def _is_blank_row(row) -> bool:
    return all(
        v is None or (isinstance(v, str) and v.strip() in ("", "\xa0"))
        for v in row
    )


# ── Filename-based fallback detection ────────────────────────────────────────
# Some services omit uur/bestemming; we infer them from the filename.
_FALLBACK_RULES: list[tuple[list[str], time, str]] = [
    # keywords (all must appear in lowercased filename)  →  uur, bestemming
    (["betekening"],             time(9, 0),  "Betekening directeur"),
    (["griffie"],                time(8, 30), "Griffie"),
]


def _detect_fallbacks(source_name: str) -> tuple[time | None, str]:
    """Return (fallback_uur, fallback_bestemming) based on filename keywords."""
    lower = source_name.lower()
    for keywords, uur, best in _FALLBACK_RULES:
        if all(kw in lower for kw in keywords):
            return uur, best
    return None, ""


def parse_dispatch(file_bytes: bytes, source_name: str = "dispatch") -> list[dict]:
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)

    fallback_uur, fallback_best = _detect_fallbacks(source_name)

    rows_out: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Find header row (contains naam + bestemming)
        header_row_idx = None
        col_map = {}
        for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
            row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
            if "naam" in row_lower and "bestemming" in row_lower:
                header_row_idx = i
                for j, val in enumerate(row):
                    lv = str(val).strip().lower() if val is not None else ""
                    if lv:
                        col_map[lv] = j
                # uur is always col 0 (may not be labelled in header)
                if "uur" not in col_map:
                    col_map["uur"] = 0
                break

        if header_row_idx is None:
            continue  # sheet has no recognizable dispatch data

        idx_uur = col_map.get("uur", 0)
        idx_cel = col_map.get("celnr", 1)
        idx_naam = col_map.get("naam", 2)
        idx_voor = col_map.get("voornaam", 3)
        idx_best = col_map.get("bestemming", 4)

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if _is_blank_row(row):
                continue

            naam_raw = row[idx_naam] if len(row) > idx_naam else None
            if naam_raw is None or str(naam_raw).strip() in ("", "\xa0"):
                continue

            uur_val  = _to_time(row[idx_uur]  if len(row) > idx_uur  else None)
            cel_val  = normalize_cell(row[idx_cel] if len(row) > idx_cel else None)
            naam_val = str(naam_raw).strip()
            voor_raw = row[idx_voor] if len(row) > idx_voor else None
            voor_val = str(voor_raw).strip() if voor_raw else None
            best_raw = row[idx_best] if len(row) > idx_best else None
            best_val = str(best_raw).strip() if best_raw else ""

            if voor_val in ("", "\xa0"):
                voor_val = None
            if best_val in ("\xa0",):
                best_val = ""

            # Apply filename-based fallbacks for missing uur / bestemming
            if uur_val is None and fallback_uur is not None:
                uur_val = fallback_uur
            if not best_val and fallback_best:
                best_val = fallback_best

            rows_out.append({
                "uur":        uur_val,
                "celnr":      cel_val,
                "naam":       naam_val,
                "voornaam":   voor_val,
                "bestemming": best_val,
                "source":     source_name,
            })

    return rows_out
