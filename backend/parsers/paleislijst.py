from __future__ import annotations
from io import BytesIO
from datetime import datetime, time
from openpyxl import load_workbook
from .normalizer import normalize_cell


_WEIGER_TERMS    = ("weigert", "weigering", "refus", "refuse")
_PALEIS_SKIP     = {"1erit", "1ste rit", "1e rit"}   # section dividers to skip
_UITHALING_HDR   = {"uithalingen", "uithalingen:"}    # marks start of uithaling section
_VIRTUAL_CELLS   = {"80000", "70000"}
_MEDISCH_TERMS   = ("medisch", "medical", "medic")


def _to_time(val) -> time | None:
    if val is None:
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    return None


def _is_zero_time(t: time | None) -> bool:
    return t is not None and t.hour == 0 and t.minute == 0


def parse_paleislijst(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Find header row
    header_row_idx = None
    col_map = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
        if "naam" in row_lower and "celnummer" in row_lower:
            header_row_idx = i
            for j, val in enumerate(row):
                if val is not None:
                    col_map[str(val).strip().lower()] = j
            break

    if header_row_idx is None:
        raise ValueError("Geen header-rij gevonden in paleislijst (verwacht: Naam, Celnummer)")

    idx_naam        = col_map.get("naam",        0)
    idx_voor        = col_map.get("voornaam",    1)
    idx_cel         = col_map.get("celnummer",   3)
    idx_onderwerp   = col_map.get("onderwerp",   4)
    idx_type        = col_map.get("type",        5)
    idx_start       = col_map.get("start",       7)
    idx_handtekening= col_map.get("handtekening",9)

    rows_out: list[dict] = []
    section = "paleis"   # switches to "uithaling" when we hit the UITHALINGEN header

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        naam_raw = row[idx_naam] if len(row) > idx_naam else None
        if naam_raw is None:
            continue

        naam_str = str(naam_raw).strip()
        if not naam_str:
            continue

        naam_lower = naam_str.lower()

        # Section header: switch to uithaling mode and skip the label row
        if naam_lower in _UITHALING_HDR:
            section = "uithaling"
            continue

        # Skip paleis divider labels (e.g. "1erit")
        if naam_lower in _PALEIS_SKIP:
            continue

        # ── Filters ──────────────────────────────────────────────────────────

        # Skip virtual / placeholder cell numbers
        cel_raw = row[idx_cel] if len(row) > idx_cel else None
        cel_str = str(cel_raw).strip() if cel_raw is not None else ""
        if cel_str in _VIRTUAL_CELLS:
            continue

        # Skip rows with no valid cell number at all
        cel_int = normalize_cell(cel_raw)
        if cel_int is None:
            continue

        # Skip 00:00 times
        start_raw = row[idx_start] if len(row) > idx_start else None
        uur_val = _to_time(start_raw)
        if _is_zero_time(uur_val):
            continue

        # Skip weigeringen
        handtekening = row[idx_handtekening] if len(row) > idx_handtekening else None
        if handtekening:
            htk_lower = str(handtekening).lower()
            if any(term in htk_lower for term in _WEIGER_TERMS):
                continue

        # ── Voornaam ─────────────────────────────────────────────────────────
        voor_raw = row[idx_voor] if len(row) > idx_voor else None
        voor_str = str(voor_raw).strip() if voor_raw else None
        if voor_str == "":
            voor_str = None

        # ── Bestemming — based on section ────────────────────────────────────
        if section == "uithaling":
            onderwerp = row[idx_onderwerp] if len(row) > idx_onderwerp else None
            type_val  = row[idx_type]      if len(row) > idx_type      else None
            combined  = " ".join(
                str(v).lower() for v in (onderwerp, type_val) if v
            )
            if any(term in combined for term in _MEDISCH_TERMS):
                bestemming = "medische uithaling"
            else:
                bestemming = "uithaling"
        else:
            bestemming = "paleis"

        rows_out.append({
            "uur":        uur_val,
            "celnr":      cel_int,
            "naam":       naam_str,
            "voornaam":   voor_str,
            "bestemming": bestemming,
            "source":     "paleislijst",
        })

    return rows_out
