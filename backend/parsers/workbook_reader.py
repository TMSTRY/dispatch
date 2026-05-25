from __future__ import annotations
"""
Unified workbook reader that handles both .xlsx (openpyxl) and .xls (xlrd).

All parsers call open_workbook(file_bytes, filename) and get back an object
with the same interface as an openpyxl Workbook:
  .active          → sheet object
  .sheetnames      → list[str]
  wb[name]         → sheet object

Each sheet object supports:
  .iter_rows(min_row=1, values_only=True) → yields tuple[Any, ...]
"""

from io import BytesIO
from datetime import date, datetime


# ── OLE2 compound document magic bytes — reliable .xls fingerprint ──────────
_XLS_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _is_xls(file_bytes: bytes, filename: str) -> bool:
    """Return True if the file is the legacy .xls format."""
    if filename:
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext == "xls":
            return True
        if ext in ("xlsx", "xlsm", "xltx", "xltm"):
            return False
    # Fall back to magic-byte detection when extension is absent / ambiguous
    return len(file_bytes) >= 8 and file_bytes[:8] == _XLS_MAGIC


# ── xlrd wrapper ─────────────────────────────────────────────────────────────

class _XlrdSheet:
    """Wraps an xlrd Sheet so it looks like an openpyxl Worksheet."""

    def __init__(self, sheet, datemode: int):
        self._sheet = sheet
        self._datemode = datemode

    def iter_rows(self, min_row: int = 1, values_only: bool = True):
        """Yield one tuple per row, starting at min_row (1-based, like openpyxl)."""
        import xlrd
        for i in range(min_row - 1, self._sheet.nrows):
            row = []
            for j in range(self._sheet.ncols):
                cell = self._sheet.cell(i, j)
                row.append(self._convert(cell, xlrd))
            yield tuple(row)

    def _convert(self, cell, xlrd):
        t = cell.ctype
        if t in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR):
            return None
        if t == xlrd.XL_CELL_TEXT:
            return cell.value
        if t == xlrd.XL_CELL_BOOLEAN:
            return bool(cell.value)
        if t == xlrd.XL_CELL_NUMBER:
            # Return int when the value is a whole number (e.g. cell numbers).
            v = cell.value
            return int(v) if v == int(v) else v
        if t == xlrd.XL_CELL_DATE:
            dt = xlrd.xldate_as_datetime(cell.value, self._datemode)
            # Pure time values have no date component (cell.value < 1).
            if cell.value < 1:
                return dt.time()
            return dt
        return cell.value


class _XlrdWorkbook:
    """Wraps an xlrd Book so it looks like an openpyxl Workbook."""

    def __init__(self, wb):
        self._wb = wb

    @property
    def sheetnames(self) -> list[str]:
        return self._wb.sheet_names()

    @property
    def active(self) -> _XlrdSheet:
        return _XlrdSheet(self._wb.sheet_by_index(0), self._wb.datemode)

    def __getitem__(self, name: str) -> _XlrdSheet:
        return _XlrdSheet(self._wb.sheet_by_name(name), self._wb.datemode)


# ── Public API ────────────────────────────────────────────────────────────────

def open_workbook(file_bytes: bytes, filename: str = ""):
    """
    Open an Excel file (either .xls or .xlsx) and return a workbook object
    with an openpyxl-compatible interface.
    """
    if _is_xls(file_bytes, filename):
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        return _XlrdWorkbook(wb)
    else:
        from openpyxl import load_workbook
        return load_workbook(filename=BytesIO(file_bytes), data_only=True)
