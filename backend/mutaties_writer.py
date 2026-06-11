from __future__ import annotations
import re
from io import BytesIO
from openpyxl import load_workbook
from rapidfuzz import process, fuzz
from parsers.mutaties_source import norm_pos, _ABBREV_MAP


def _short_name(name: str) -> str:
    """Format 'LASTNAME Firstname' → 'LASTNAME F' (no period)."""
    tokens = name.strip().split()
    if len(tokens) <= 1:
        return name
    # Find the first token that is NOT all-caps → that's the first name
    for i, token in enumerate(tokens):
        if not token.isupper():
            lastname = " ".join(tokens[:i])
            initial = token[0].upper()
            return f"{lastname} {initial}" if lastname else initial
    return name  # all tokens uppercase → no first name found

# ── Nacht role detection ──────────────────────────────────────────────────────
# These patterns identify nacht role labels in the template's col H.
# They are distinct from other col H content (WANDELING, globe, rondeweg, etc.)

_NACHT_PATTERNS = [
    re.compile(r"^postoverste$", re.I),
    re.compile(r"^pba nacht", re.I),
    re.compile(r"^pba vleugel", re.I),
    re.compile(r"^pba receptie$", re.I),
    re.compile(r"^pba ziekenhuis", re.I),
]


def _is_nacht_label(s: str) -> bool:
    return any(p.match(s.strip()) for p in _NACHT_PATTERNS)


# ── Matching helpers ──────────────────────────────────────────────────────────

def _match_pos(label: str, keys: list[str]) -> str | None:
    """Fuzzy-match a template position label to a source position key."""
    norm = norm_pos(label)
    if not keys:
        return None
    r = process.extractOne(norm, keys, scorer=fuzz.token_set_ratio, score_cutoff=70)
    return r[0] if r else None


def _match_nacht(label: str, nacht_keys: list[str]) -> str | None:
    """Fuzzy-match a template nacht label to a source nacht role key."""
    norm = label.strip().lower()
    if not nacht_keys:
        return None
    r = process.extractOne(norm, nacht_keys, scorer=fuzz.token_set_ratio, score_cutoff=60)
    return r[0] if r else None


# ── Main filler ───────────────────────────────────────────────────────────────

def fill_mutaties_template(template_bytes: bytes, roster: dict) -> bytes:
    """
    Fill the mutatielijst template (weekdag or weekend sheet) with roster data.

    Template structure (row 3+):
      Col A (idx 0): position label for vroeg/dag/laat
      Col B (idx 1): vroeg person → FILL
      Col D (idx 3): dag person   → FILL
      Col F (idx 5): laat person  → FILL
      Col H (idx 7): nacht role label (even rows) → next row gets the nacht person → FILL

    Nacht convention confirmed by user: label on row N, person name on row N+1.
    Each position may span 2 rows (two persons per shift).
    """
    wb = load_workbook(BytesIO(template_bytes))

    is_weekend = roster.get("is_weekend", False)
    sheet_name = "weekend" if is_weekend else "weekdag"
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    shifts    = roster.get("shifts", {})
    nacht     = roster.get("nacht",  {})
    pos_keys  = list(shifts.keys())
    nacht_keys = list(nacht.keys())

    last_pos_norm: str | None = None
    pos_row_count = 0
    pending_nacht_label: str | None = None   # nacht label seen on the previous row

    for row in ws.iter_rows(min_row=3, values_only=False):
        col_a = row[0].value
        col_h = row[7].value

        # ── Left section: vroeg / dag / laat ─────────────────────────────────
        if col_a is not None and str(col_a).strip():
            pos_label = str(col_a).strip()
            pos_norm  = norm_pos(pos_label)

            if pos_norm == last_pos_norm:
                pos_row_count += 1     # second row for same position → person 2
            else:
                last_pos_norm  = pos_norm
                pos_row_count  = 1     # first row for this position → person 1

            matched = _match_pos(pos_label, pos_keys)
            if matched:
                data = shifts[matched]
                pi   = pos_row_count - 1   # 0 = first person, 1 = second

                vroeg = data.get("vroeg", [])
                dag   = data.get("dag",   [])
                laat  = data.get("laat",  [])

                if pi < len(vroeg): row[1].value = _short_name(vroeg[pi])   # col B
                if pi < len(dag):   row[3].value = _short_name(dag[pi])     # col D
                if pi < len(laat):  row[5].value = _short_name(laat[pi])    # col F

        # ── Right section: nacht ──────────────────────────────────────────────
        # User confirmed: label on row N, person name goes in col H of row N+1.
        if pending_nacht_label is not None:
            mk = _match_nacht(pending_nacht_label, nacht_keys)
            if mk and nacht.get(mk):
                row[7].value = _short_name(nacht[mk])
            pending_nacht_label = None

        # Queue this row's nacht label for the next iteration
        if col_h is not None and _is_nacht_label(str(col_h)):
            pending_nacht_label = str(col_h).strip()

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
