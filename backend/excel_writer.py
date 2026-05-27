from __future__ import annotations
from datetime import date, time, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sorter import insert_hour_separators


# ── Formatting constants (from golden standard analysis) ─────────────────────

_FONT_TITLE = Font(name="Arial", size=14, bold=True)
_FONT_HEADER_MAIN = Font(name="Arial", size=10, bold=True)
_FONT_HEADER_SMALL = Font(name="Arial", size=8, bold=True)
_FONT_DATA = Font(name="Calibri", size=10)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_HEADER_SMALL = Alignment(horizontal="center")

_SIDE_MED = Side(border_style="medium")
_SIDE_THIN = Side(border_style="thin")
_SIDE_NONE = Side(border_style=None)

_BORDER_TITLE = Border(
    left=_SIDE_MED, right=_SIDE_MED, top=_SIDE_MED, bottom=_SIDE_MED
)
_BORDER_DATE = Border(
    left=_SIDE_NONE, right=_SIDE_MED, top=_SIDE_MED, bottom=_SIDE_MED
)
_BORDER_HDR_FIRST = Border(
    left=_SIDE_MED, right=_SIDE_THIN, top=_SIDE_MED, bottom=_SIDE_NONE
)
_BORDER_HDR_MID = Border(
    left=_SIDE_THIN, right=_SIDE_THIN, top=_SIDE_MED, bottom=_SIDE_NONE
)
_BORDER_HDR_LAST = Border(
    left=_SIDE_THIN, right=_SIDE_MED, top=_SIDE_MED, bottom=_SIDE_NONE
)
_BORDER_DATA = Border(
    left=_SIDE_THIN, right=_SIDE_THIN, top=_SIDE_THIN, bottom=_SIDE_THIN
)
_BORDER_DATA_FIRST = Border(
    left=_SIDE_THIN, right=_SIDE_THIN, top=_SIDE_THIN, bottom=_SIDE_THIN
)
_BORDER_DATA_LAST = Border(
    left=_SIDE_THIN, right=_SIDE_THIN, top=_SIDE_THIN, bottom=_SIDE_THIN
)

_ROW_HEIGHT_TITLE = 31.5
_ROW_HEIGHT_HEADER = 60.0
_ROW_HEIGHT_DATA = 19.5

_COL_WIDTHS = {
    "A": 6.42578125,
    "B": 5.85546875,
    "C": 21.0,
    "D": 14.0,
    "E": 23.7109375,
    "F": 5.7109375,
    "G": 7.0,
    "H": 5.5,
}

_DUTCH_DAYS = ["Maandag", "Dinsdag", "Woensdag", "Donderdag", "Vrijdag", "Zaterdag", "Zondag"]
_DUTCH_MONTHS = [
    "Januari", "Februari", "Maart", "April", "Mei", "Juni",
    "Juli", "Augustus", "September", "Oktober", "November", "December",
]


def dutch_date_filename(d: date) -> str:
    day_name = _DUTCH_DAYS[d.weekday()]
    month_name = _DUTCH_MONTHS[d.month - 1]
    return f"{day_name} {d.day:02d} {month_name} {d.year}.xlsx"


def _date_display(d: date) -> str:
    return f"{d.day:02d}/{d.month:02d}/{str(d.year)[2:]}"


def _apply_col_widths(ws):
    for col, width in _COL_WIDTHS.items():
        ws.column_dimensions[col].width = width


def _write_header(ws, title: str, disp_date: str, is_lijst_disp: bool = False):
    """Write rows 1 and 2: title + date merged header, then column labels."""
    # Row 1
    ws.row_dimensions[1].height = _ROW_HEIGHT_TITLE
    ws.merge_cells("A1:E1")
    ws.merge_cells("F1:H1")

    c_title = ws["A1"]
    c_title.value = title
    c_title.font = _FONT_TITLE
    c_title.alignment = _ALIGN_CENTER
    c_title.border = _BORDER_TITLE

    c_date = ws["F1"]
    c_date.value = disp_date
    c_date.font = _FONT_TITLE
    c_date.alignment = _ALIGN_CENTER
    c_date.border = _BORDER_DATE

    if is_lijst_disp:
        ws.column_dimensions["J"].width = 82.140625
        c_note = ws["J1"]
        c_note.value = "Geeft hier de juiste datum in"
        c_note.font = Font(name="Arial", size=16)
        c_note.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Row 2 — column headers
    ws.row_dimensions[2].height = _ROW_HEIGHT_HEADER
    headers = [
        ("A2", "uur", _FONT_HEADER_MAIN, _BORDER_HDR_FIRST, "h:mm;@"),
        ("B2", "celnr", _FONT_HEADER_MAIN, _BORDER_HDR_MID, "General"),
        ("C2", "naam", _FONT_HEADER_MAIN, _BORDER_HDR_MID, "General"),
        ("D2", "voornaam", _FONT_HEADER_MAIN, _BORDER_HDR_MID, "General"),
        ("E2", "bestemming", _FONT_HEADER_MAIN, _BORDER_HDR_MID, "General"),
        ("F2", "opgeroepen", _FONT_HEADER_SMALL, _BORDER_HDR_MID, "General"),
        ("G2", "doorgestuurd", _FONT_HEADER_SMALL, _BORDER_HDR_MID, "General"),
        ("H2", "terug", _FONT_HEADER_SMALL, _BORDER_HDR_LAST, "General"),
    ]
    for coord, val, font, border, nfmt in headers:
        cell = ws[coord]
        cell.value = val
        cell.font = font
        cell.alignment = _ALIGN_CENTER if font == _FONT_HEADER_MAIN else _ALIGN_HEADER_SMALL
        cell.border = border
        cell.number_format = nfmt


def _write_data_row(ws, row_num: int, row: dict | None):
    """Write one data row (None = blank separator row)."""
    ws.row_dimensions[row_num].height = _ROW_HEIGHT_DATA

    cells = [
        ws.cell(row=row_num, column=c) for c in range(1, 9)
    ]
    for cell in cells:
        cell.border = _BORDER_DATA
        cell.alignment = _ALIGN_CENTER
        cell.font = _FONT_DATA

    if row is None:
        return  # blank separator with borders applied

    uur = row.get("uur")
    celnr = row.get("celnr")
    naam = row.get("naam", "")
    voornaam = row.get("voornaam") or ""
    # Only the first given name on the dispatch list
    voornaam = voornaam.split()[0] if voornaam.strip() else ""
    bestemming = row.get("bestemming", "")

    cells[0].value = uur if isinstance(uur, time) else None
    cells[0].number_format = "h:mm"
    cells[1].value = celnr
    cells[2].value = naam
    cells[3].value = voornaam
    cells[4].value = bestemming


def _write_sheet(ws, title: str, disp_date: str, rows: list, is_lijst_disp=False, add_hour_separators=False):
    _apply_col_widths(ws)
    _write_header(ws, title, disp_date, is_lijst_disp=is_lijst_disp)

    display_rows = insert_hour_separators(rows) if add_hour_separators else rows

    for i, row in enumerate(display_rows, start=3):
        _write_data_row(ws, i, row)


def generate_workbook(tabs: dict[str, list[dict]], target_date: date) -> Workbook:
    wb = Workbook()
    disp_date = _date_display(target_date)

    # Remove default sheet
    wb.remove(wb.active)

    # Tab order: lijst disp, Verzorging, sectie 1..10, sectie 11
    sheet_defs = [
        ("lijst disp", "DISPATCHER", True, False),
        ("verzorging", "VERZORGING", False, True),  # key must match build_tabs dict (lowercase)
    ] + [
        (f"sectie {i}", f"SECTIE {i}", False, False) for i in range(1, 12)
    ]

    for tab_key, title, is_ld, add_sep in sheet_defs:
        ws = wb.create_sheet(title=tab_key)
        rows = tabs.get(tab_key, [])
        _write_sheet(ws, title, disp_date, rows, is_lijst_disp=is_ld, add_hour_separators=add_sep)

    return wb
