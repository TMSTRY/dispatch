from __future__ import annotations
from io import BytesIO
from openpyxl import load_workbook


_SECTION_SHEETS = {
    "1": "1ste", "2": "2de", "3": "3de", "4": "4de", "5": "5de",
    "6": "6de",  "7": "7de", "8": "8ste", "9": "9de",
}


def _norm(name) -> str:
    return " ".join(str(name).upper().split()) if name is not None else ""


def _section(cel) -> str | None:
    s = str(cel).strip() if cel is not None else ""
    return _SECTION_SHEETS.get(s[0]) if s and s[0].isdigit() else None


def _sort_key(row):
    try:
        return (int(row[0]), "")
    except (TypeError, ValueError):
        return (9999, str(row[0]) if row[0] else "")


def update_werkers(werkers_bytes: bytes, mutaties: dict) -> tuple[bytes, dict]:
    """
    Apply mutations to Werkers.xlsx:
      1. Update cell numbers (MUTATIE VAN → NAAR), matched by name
      2. Remove UITGAANDEN (skip reden == PV)
      3. Sort by cell number
      4. Populate section sheets (1ste … 9de)
    Returns (updated_bytes, summary).
    """
    wb = load_workbook(BytesIO(werkers_bytes))

    if "Totaal Dispatcher" not in wb.sheetnames:
        raise ValueError("Geen 'Totaal Dispatcher' sheet gevonden.")

    ws = wb["Totaal Dispatcher"]
    summary: dict = {"removed": [], "updated": []}

    # Read all data rows (skip header at row 1)
    rows: list[list] = []
    for i, row_vals in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if any(v is not None for v in row_vals):
            rows.append(list(row_vals))

    # ── 1. Apply cell mutations ───────────────────────────────────────────────
    for mut in mutaties.get("mutaties", []):
        target = _norm(mut["naam"])
        van    = str(mut["van"]).strip()
        naar   = mut["naar"]
        for row in rows:
            if _norm(row[1]) == target and str(row[0]).strip() == van:
                summary["updated"].append({"naam": row[1], "van": row[0], "naar": naar})
                row[0] = naar
                break  # one match per mutatie row

    # ── 2. Remove uitgaanden (except PV) ─────────────────────────────────────
    to_remove = {
        _norm(u["naam"])
        for u in mutaties.get("uitgaanden", [])
        if str(u.get("reden", "")).upper().strip() != "PV"
    }
    new_rows: list[list] = []
    for row in rows:
        if _norm(row[1]) in to_remove:
            summary["removed"].append({"naam": row[1], "cel": row[0]})
        else:
            new_rows.append(row)

    # ── 3. Sort by cell number ────────────────────────────────────────────────
    new_rows.sort(key=_sort_key)

    # ── 4. Write Totaal Dispatcher ────────────────────────────────────────────
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    for i, row_data in enumerate(new_rows, 2):
        for j, val in enumerate(row_data):
            ws.cell(row=i, column=j + 1, value=val)

    # ── 5. Populate section sheets ────────────────────────────────────────────
    for sheet_name in _SECTION_SHEETS.values():
        if sheet_name not in wb.sheetnames:
            continue
        ws_sec = wb[sheet_name]
        for row in ws_sec.iter_rows(min_row=2, max_row=ws_sec.max_row):
            for cell in row:
                cell.value = None
        sec_rows = [r for r in new_rows if _section(r[0]) == sheet_name]
        for i, row_data in enumerate(sec_rows, 2):
            for j, val in enumerate(row_data[:6]):
                ws_sec.cell(row=i, column=j + 1, value=val)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), summary
